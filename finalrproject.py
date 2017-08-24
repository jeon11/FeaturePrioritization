import openpyxl 
import numpy as np
import matplotlib.mlab as mlab
import matplotlib.pyplot as plt
from sklearn import linear_model
from sklearn.linear_model import LogisticRegression

wb = openpyxl.load_workbook('data.xlsx')
sheet = wb.get_sheet_by_name('conversions')

ob1 = []
for i in range(2,27):
	ob1.append(sheet.cell(row=i, column=1).value)
ob2 = []
for i in range(2,29):
	ob2.append(sheet.cell(row=i, column=3).value)
ob3 = []
for i in range(2,22):
	ob3.append(sheet.cell(row=i, column=5).value)
ob4 = []
for i in range(2,19):
	ob4.append(sheet.cell(row=i, column=7).value)
ob5 = []
for i in range(2,24):
	ob5.append(sheet.cell(row=i, column=9).value)

total = len(ob1)+len(ob2)+len(ob3)+len(ob4)+len(ob5)

s_single = 0; c_single = 0; o_single = 0; #single most indicator
s = 0; c = 0; o = 0; sc = 0; so = 0; co = 0; sco = 0; na = 0 #combination of features

for i in ob1: #data for item 1
	if i == 1.1:
		s_single += 1; s += 1
	if i == 1.2:
		s_single += 1; c_single += 1; sc += 1
	if i == 1.3:
		s_single += 1; o_single += 1; so += 1
	if i == 1.4:
		s_single += 1; c_single += 1; o_single += 1; sco += 1
	if i == 1.5:
		na += 1
	if i == 1.6:
		c_single += 1; c += 1
	if i == 1.7:
		o_single += 1; o += 1
	if i == 1.8:
		c_single += 1; o_single += 1; co += 1

for j in ob2: #data for item 2
	if j == 2.1:
		s_single += 1; o_single += 1; so += 1
	if j == 2.2:
		s_single += 1; c_single += 1; o_single += 1; sco += 1
	if j == 2.3:
		s_single += 1; s += 1
	if j == 2.4:
		s_single += 1; c_single += 1; sc += 1
	if j == 2.5:
		o_single += 1; o += 1
	if j == 2.6:
		c_single += 1; o_single += 1; co += 1
	if j == 2.7:
		na += 1
	if j == 2.8:
		c_single += 1; c += 1

for k in ob3: #data for item 3
	if k == 3.1:
		s_single += 1; o_single += 1; so += 1
	if k == 3.2:
		s_single += 1; c_single += 1; o_single += 1; sco += 1
	if k == 3.3:
		o_single += 1; o += 1
	if k == 3.4:
		c_single += 1; o_single += 1; co += 1
	if k == 3.5:
		na += 1
	if k == 3.6:
		c_single += 1; c += 1
	if k == 3.7:
		s_single += 1; s += 1
	if k == 3.8:
		s_single += 1; c_single += 1; sc += 1

for l in ob4: #data for item 4
	if l == 4.1:
		s_single += 1; c_single += 1; sc += 1
	if l == 4.2:
		s_single += 1; s += 1
	if l == 4.3:
		s_single += 1; c_single += 1; o_single += 1; sco += 1
	if l == 4.4:
		s_single += 1; o_single += 1; so += 1
	if l == 4.5:
		o_single += 1; o += 1
	if l == 4.6:
		c_single += 1; o_single += 1; co += 1
	if l == 4.7:
		na += 1
	if l == 4.8:
		c_single += 1; c += 1

for m in ob5: #data for item 5
	if m == 5.1:
		s_single += 1; o_single += 1; so += 1
	if m == 5.2:
		s_single += 1; s += 1
	if m == 5.3:
		o_single += 1; o += 1
	if m == 5.4:
		na += 1
	if m == 5.5:
		s_single += 1; c_single += 1; o_single += 1; sco += 1
	if m == 5.6:
		s_single += 1; c_single += 1; sc += 1
	if m == 5.7:
		c_single += 1; o_single += 1; co += 1
	if m == 5.8:
		c_single += 1; c += 1

x = ('size', 'color', 'orientation')
y_pos = np.arange(len(x))
y = [s_single, c_single, o_single]
 
plt.bar(y_pos, y, align='center', alpha=0.5)
plt.xticks(y_pos, x)
plt.xlabel('variables')
plt.ylabel('number of "in category"')
plt.title('Single most effective variable in categorization')
# plt.show()

x = ('size', 'color', 'orient', 'size&color', 'size&orient', 'color&orient', 'all', 'none')
y_pos = np.arange(len(x))
y = [s, c, o, sc, so, co, sco, na]
 
plt.bar(y_pos, y, align='center', alpha=0.5)
plt.xticks(y_pos, x, rotation='15')
plt.xlabel('feature combinations')
plt.ylabel('number of "in category"')
plt.title('Combinations and categorization')
# plt.show()

# references
# http://www.wikihow.com/Assess-Statistical-Significance
# http://stackoverflow.com/questions/22306341/python-sklearn-how-to-calculate-p-values
# http://scikit-learn.org/stable/modules/generated/sklearn.feature_selection.f_regression.html
# http://stackoverflow.com/questions/27928275/find-p-value-significance-in-scikit-learn-linearregression
# https://onlinecourses.science.psu.edu/stat504/node/150
# http://scikit-learn.org/stable/auto_examples/linear_model/plot_logistic.html#sphx-glr-auto-examples-linear-model-plot-logistic-py

# model = LogisticRegression(fit_intercept=False, C = 1e9)
# mdl = model.fit(y)

# this is our test set, it's just a straight line with some
# Gaussian noise
xmin, xmax = -5, 5
n_samples = 117
np.random.seed(0)

# X = np.random.normal(size=n_samples) #change this to S, C, O by i.e. 1, 1, 1, 0, 0, etc
S = [1,1,0,0,1,0,1,1,1,1,1,0,1,1,1,0,1,1,0,0,1,1,0,1,0,1,0,1,1,1,1,1,0,1,1,0,0,1,1,1,1,0,0,1,0,1,1,1,1,0,0,1,1,0,0,1,0,1,0,1,1,1,0,1,1,0,0,0,0,1,1,1,1,1,1,1,1,1,0,0,1,1,1,1,0,0,1,1,0,1,1,1,0,0,1,1,1,1,0,1,1,1,1,1,1,1,0,0,1,0,1,0,1,0,1,0,1]
C = [1,1,1,1,1,1,0,1,0,1,1,1,1,0, 1,1,1,0,1,0,1,0,1,0,1,1,1,0,1,0,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,1,0,1,0,1,1,1,1,1,1,1,1,1,0,1,0,1,0,1,0,1,0,1,0,1,1,1,0,0,1,1,1,1,1,1,1,1,1,1,1,1,0,0,1,0,0,1,1,1,1,1,1,1,1,1,1,1,1,0,0,1,1,0,0,1,1,1]
O = [0,1, 0, 1, 1, 0, 0, 0, 1, 1, 1, 0, 1, 0, 1, 1, 1, 1, 1, 1, 1, 1, 1, 0, 0, 0, 0, 1, 1, 0, 0, 1, 1, 1, 0, 1, 0, 1, 0, 1, 0, 1, 0, 1, 1, 1, 0, 1, 0, 1, 0, 1, 1, 1, 1, 1, 1, 1, 0, 0, 1, 1, 1, 1, 1, 1, 1, 0, 0, 0, 0, 1, 1, 0, 0, 1, 1, 1, 1, 1, 0, 1, 0, 1, 1, 0, 0, 1, 0, 1, 1, 1, 1, 1, 1, 0, 1, 0, 1, 0, 1, 0, 1, 0, 1, 0, 1, 0, 1, 1, 1, 1, 1, 1, 1, 1, 1]

S = np.asarray(S).astype(float)
y = (S > 0).astype(np.float) #range from 0 to 1
S += .2 * np.random.normal(size=n_samples) #random noise 

S = S[:, np.newaxis]
# run the classifier
clf = linear_model.LogisticRegression(C=1e5)
clf.fit(S, y)
Scoef = clf.coef_
# and plot the result
plt.figure(1, figsize=(4, 3))
plt.clf()
plt.scatter(S.ravel(), y, color='black', zorder=20)
S_test = np.linspace(-1, 2, 300)


def model(x):
    return 1 / (1 + np.exp(-x))
loss = model(S_test * clf.coef_ + clf.intercept_).ravel()
plt.plot(S_test, loss, color='red', linewidth=3)

plt.axhline(.5, color='.5')
plt.title('Logistic regression for feature size')
plt.ylabel('y')
plt.xlabel('X')
plt.xticks(range(-1, 2))
plt.yticks([0, 0.5, 1])
plt.ylim(-.25, 1.25)
plt.xlim(-1, 2)
plt.legend(('Logistic Regression Model', 'axis'),
           loc="lower right", fontsize='small')
plt.show()

# %%%%%%%%%%%%%%%%%%%%%%%%%%%%
C = np.asarray(C).astype(float)
y = (C > 0).astype(np.float) #range from 0 to 1
C += .2 * np.random.normal(size=n_samples) #random noise 

C = C[:, np.newaxis]
# run the classifier
clf = linear_model.LogisticRegression(C=1e5)
clf.fit(C, y)
Ccoef = clf.coef_
# and plot the result
plt.figure(1, figsize=(4, 3))
plt.clf()
plt.scatter(C.ravel(), y, color='black', zorder=20)
C_test = np.linspace(-1, 2, 300)

loss = model(C_test * clf.coef_ + clf.intercept_).ravel()
plt.plot(C_test, loss, color='blue', linewidth=3)
plt.axhline(.5, color='.5')

plt.title('Logistic regression for feature color')
plt.ylabel('y')
plt.xlabel('X')
plt.xticks(range(-1, 2))
plt.yticks([0, 0.5, 1])
plt.ylim(-.25, 1.25)
plt.xlim(-1, 2)
plt.legend(('Logistic Regression Model', 'axis'),
           loc="lower right", fontsize='small')
plt.show()

# %%%%%%%%%%%%%%
O = np.asarray(O).astype(float)
y = (O > 0).astype(np.float) #range from 0 to 1
O += .2 * np.random.normal(size=n_samples) #random noise 

O = O[:, np.newaxis]
# run the classifier
clf = linear_model.LogisticRegression(C=1e5)
clf.fit(O, y)
Ocoef = clf.coef_
# and plot the result
plt.figure(1, figsize=(4, 3))
plt.clf()
plt.scatter(O.ravel(), y, color='black', zorder=20)
O_test = np.linspace(-1, 2, 300)

loss = model(S_test * clf.coef_ + clf.intercept_).ravel()
plt.plot(S_test, loss, color='green', linewidth=3)

plt.axhline(.5, color='.5')

plt.title('Logistic regression for feature orientation')
plt.ylabel('y')
plt.xlabel('X')
plt.xticks(range(-1, 2))
plt.yticks([0, 0.5, 1])
plt.ylim(-.25, 1.25)
plt.xlim(-1, 2)
plt.legend(('Logistic Regression Model', 'axis'),
           loc="lower right", fontsize='small')
plt.show()
